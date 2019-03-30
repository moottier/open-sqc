import os
import re
import statistics
from datetime import datetime
from typing import List, Tuple, Union, Any, Generator, Type

import matplotlib.ticker as mticker
import openpyxl
import openpyxl.cell
import openpyxl.utils.exceptions
from openpyxl.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

from ccrev import config
from ccrev.charts.charting_base import ControlChart
from ccrev.reporting import Report
from ccrev.rule_checking import RuleChecker


class DataExtractor:
    def __init__(self):
        pass

    @staticmethod
    def get_excel_data(
            src_file, min_row=None, max_row=None, min_col=None, max_col=None, worksheet_index=None,
            read_only=True, data_only=True, values_only=True
    ) -> Union[List[Any], Any]:
        # TODO
        # make this a generator not a function
        # evaluating every entry before returning takes too much time

        wb = DataExtractor.load_workbook(
                src_file,
                read_only=read_only,
                data_only=data_only
        )
        ws = DataExtractor.load_worksheet(
                wb,
                worksheet_index
        )
        data_iter = DataExtractor._get_data_iter(
                ws,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=values_only
        )

        if min_col is max_col and min_row is max_row:
            # data from single cell
            data = [val[0] for val in data_iter if val[0] is not None][0]
        elif min_col is max_col:
            # data from single column
            data = [val[0] for val in data_iter if val[0] is not None]
        else:
            # data from contiguous columns
            data = [vals for vals in data_iter if any(val for val in vals is not None)]

        try:
            if any(val is '#REF!' for val in data):
                # TODO need an error handling module for all these excel errors
                print('Bad excel reference loaded.')
        except TypeError:
            if data is '#REF!':  # TODO put this somewhere else
                print('Bad excel reference loaded.')
        return data

    @staticmethod
    def gen_files(src_dir: str, include_extensions: Union[str, Tuple[str]] = None, excludes: Tuple[str] = None) -> str:
        for file in os.listdir(src_dir):
            file = os.path.join(src_dir, file)
            if include_extensions and not file.endswith(include_extensions):
                continue
            if excludes and any(exclude in file for exclude in excludes):
                continue
            else:
                yield file

    @staticmethod
    def remove_file_extensions(s: str) -> str:
        return re.sub(r'\..+$', '', s)

    @staticmethod
    def load_workbook(file, read_only=True, data_only=True) -> Workbook:
        return openpyxl.load_workbook(file, read_only=read_only, data_only=data_only)

    @staticmethod
    def load_worksheet(workbook: Workbook, sheet_index: int) -> Union[Worksheet, ReadOnlyWorksheet]:
        return workbook.worksheets[sheet_index]

    @staticmethod
    def _get_data_iter(
            worksheet: Union[Worksheet, ReadOnlyWorksheet],
            min_col: int, max_col: int, min_row: int, max_row: int = None, values_only=True
    ) -> Generator[Any, None, None]:
        return worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=values_only
        )

    @staticmethod
    def clean_file_names(src_file):
        src_file = os.path.basename(src_file)
        src_file = DataExtractor.remove_file_extensions(src_file)
        return src_file


class Reviewer:
    DefaultReport: Type[Report] = Report

    def __init__(self, data_col=None, index_col=None, min_row=None, max_row=None, rules=None,
                 data_sheet_index=None, load_stats_from_src=False, **stats_data_addresses):

        self.data_col = data_col
        self.index_col = index_col
        self.min_row = min_row
        self.max_row = max_row
        self.stats_data_addresses = stats_data_addresses
        self.data_sheet_index = data_sheet_index

        self.report: Report = None
        self.data_extractor: DataExtractor = DataExtractor()
        self.rule_checker: RuleChecker = RuleChecker(rules=rules)

        self.config = {
            'try_to_load_stats_data'   : load_stats_from_src,
        }

        # List[List[title, ControlChart]]
        self.control_chart_data: List[List[Union[str, ControlChart]]] = []
        self._active_data = None

    @property
    def control_charts(self) -> List[ControlChart]:
        return [chart_item[1] for chart_item in self.control_chart_data]

    @property
    def chart_src_files(self) -> List[str]:
        return [chart_item[0] for chart_item in self.control_chart_data]

    @property
    def chart_titles(self):
        return [chart_item[1].title for chart_item in self.control_chart_data]

    def add_control_charts_from_directory(self, src_dir: str,
                                          chart_type: Union[Type[ControlChart], List[Type[ControlChart]]],
                                          file_types=config.EXCEL_FILE_EXTENSIONS,
                                          ignore=config.IGNORE_FILES) -> None:

        files = self.data_extractor.gen_files(src_dir, file_types, ignore)

        try:
            for chart_type, file in zip(chart_type, files):
                title = self.data_extractor.clean_file_names(file)
                self.add_chart_from_file(file, chart_type=chart_type, title=title)

        except TypeError:
            for file in files:
                title = self.data_extractor.clean_file_names(file)
                self.add_chart_from_file(file, chart_type=chart_type, title=title)

    def load_chart_data(self, *, src_file: str = None, chart_title: str = None, chart: ControlChart = None) -> None:

        if not chart and src_file:
            chart = self.control_charts[self.chart_src_files.index(src_file)]

        if not chart and chart_title:
            chart = self.control_charts[self.chart_titles.index(chart_title)]

        if not chart:
            raise ValueError('Chart not found!')

        chart_index = self.control_charts.index(chart)
        src_file = self.chart_src_files[chart_index]

        # TODO
        #  return all columns from a single function call
        #  maybe something like self._load_data() <- self._load_cols() <- openpyxl.Workbook.col_iter()
        #  or a loader class that stacks and processes load request
        data = self._load_data(src_file)
        data_index = self._load_index(src_file)

        if len(data) is not len(data_index):
            # TODO log this? warn for this? print is okay for now but could improve...
            print('%s: data and data-index lengths mismatched' % src_file)

        if self.config['try_to_load_stats_data']:
            # TODO shouldn't need try-except here for this problem
            # indexerror is being raised when file formatting doesn't match config
            try:
                stdev = self._load_st_dev(src_file)
            except IndexError:
                stdev = statistics.stdev(data)
                print('STDEV could not be loaded for %s. Calculating instead' % src_file)
            try:
                mean = self._load_mean(src_file)
            except IndexError:
                mean = statistics.mean(data)
                print('MEAN could not be loaded for %s. Calculating instead' % src_file)
        else:
            stdev, mean = statistics.stdev(data), statistics.mean(data)

        chart.y_data = data
        chart.x_data = data_index
        chart.stdev = stdev
        chart.mean = mean

    def load_all_data(self) -> None:
        for chart in self.control_charts:
            self.load_chart_data(chart=chart)
            chart.make_plot()

    def add_chart_from_file(self, src_file: str, chart_type: Type[ControlChart], title=None) -> None:
        title = title if title else self.data_extractor.clean_file_names(src_file)
        if title in self.chart_titles:
            raise ValueError('Chart title cannot be a duplicate')
        chart = chart_type(
                x_data=None,
                y_data=None,
                signals=None,
                title=title,
        )
        self.control_chart_data.append([src_file, chart])

    def _load_data(self, src_file: str) -> List:
        return self.data_extractor.get_excel_data(
                src_file,
                min_row=self.min_row,
                max_row=self.max_row,
                min_col=self.data_col,
                max_col=self.data_col,
                worksheet_index=self.data_sheet_index
        )

    def _load_index(self, src_file: str) -> List:
        data = self.data_extractor.get_excel_data(
                src_file,
                min_row=self.min_row,
                max_row=self.max_row,
                min_col=self.index_col,
                max_col=self.index_col,
                worksheet_index=self.data_sheet_index
        )
        if not all(isinstance(val, datetime) for val in data):
            # TODO more than just datetime should be acceptable as index in future
            print('%s: Non-datetime index found' % src_file)
        return data

    def _load_st_dev(self, src_file: str) -> float:
        return self.data_extractor.get_excel_data(
                src_file,
                min_row=self.stats_data_addresses[config.STDEV][0],
                max_row=self.stats_data_addresses[config.STDEV][0],
                min_col=self.stats_data_addresses[config.STDEV][1],
                max_col=self.stats_data_addresses[config.STDEV][1],
                worksheet_index=self.data_sheet_index
        )

    def _load_mean(self, src_file: str) -> float:
        return self.data_extractor.get_excel_data(
                src_file,
                min_row=self.stats_data_addresses[config.MEAN][0],
                max_row=self.stats_data_addresses[config.MEAN][0],
                min_col=self.stats_data_addresses[config.MEAN][1],
                max_col=self.stats_data_addresses[config.MEAN][1],
                worksheet_index=self.data_sheet_index
        )

    @staticmethod
    def _calculate_st_dev(data: List[float]) -> float:
        return statistics.stdev(data)

    @staticmethod
    def _calculate_mean(data: List[float]) -> float:
        return statistics.mean(data)

    def check_all_rules(self):
        for chart in self.control_charts:
            if not chart.plotted_x_data:
                print(f'Trying to check chart without loading data: {chart.title}')
                continue

            chart.signals = self.rule_checker.check_all_rules(
                    chart.plotted_y_data,
                    st_dev=chart.stdev,
                    mean=chart.mean
            )

    def build_report(self, report_name=None, save=True):
        self.report = Reviewer.DefaultReport()
        for chart in self.control_charts:
            self.report.add_chart(chart, signal_labels=chart.x_data)
        self.report.name = report_name

        if save:
            self.save_report()

    def save_report(self):
        self.report.save()

    def swap_chart_order(self, pos1, pos2) -> None:
        self.control_chart_data[pos2], self.control_chart_data[pos1] = \
            self.control_chart_data[pos1], self.control_chart_data[pos2]

    def move_chart_up(self, pos) -> None:
        Reviewer._can_move(pos, pos - 1) and self.swap_chart_order(pos, pos - 1)

    def move_chart_down(self, pos) -> None:
        Reviewer._can_move(pos, pos + 1) and self.swap_chart_order(pos, pos + 1)

    @staticmethod
    def _can_move(pos1, pos2) -> bool:
        args = [pos1, pos2]
        if any(arg < 0 for arg in args):
            raise IndexError('list index out of range')
        return True

    def overwrite_mean(self, chart_title, mean):
        chart_idx = self.chart_titles.index(chart_title)
        self.control_charts[chart_idx].mean = mean

    def overwrite_stdev(self, chart_title, stdev):
        chart_idx = self.chart_titles.index(chart_title)
        self.control_charts[chart_idx].stdev = stdev

    def resize_chart_axes(self, chart_title, x_min, x_max, y_min, y_max):
        chart_idx = self.chart_titles.index(chart_title)
        self.control_charts[chart_idx].plot.resize_plot_axes(x_min, x_max, y_min, y_max)

    def label_x_axis(self, chart_title):
        chart_idx = self.chart_titles.index(chart_title)
        chart = self.control_charts[chart_idx]
        x_axis = chart.plot.axes.xaxis
        x_axis.set_major_formatter(mticker.IndexFormatter(
                [f'{val.month}/{val.day}' if isinstance(val, datetime) else val for val in chart.x_data]
        ))
