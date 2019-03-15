import datetime
import itertools
import os
from typing import List, Dict, Tuple, Generator, Union, Any

import openpyxl
import openpyxl.cell
import openpyxl.worksheet.table
from openpyxl.cell import Cell
from openpyxl.utils.cell import range_boundaries

PATH = r'H:\code\ccleaner\Individuals Chart.xltx'
WORKSHEET_INDEX = 0

HEADER_VALUES = {
    (1, 1): 'Date',
    (1, 2): 'Time',
    (1, 3): 'Operator',
    (1, 4): 'Measured Value',
    (1, 5): 'Mean Value',
    (1, 6): 'Datetime',
    (1, 7): '+1SL',
    (1, 8): '-1SL',
    (1, 9): 'UWL',
    (1, 10): 'LWL',
    (1, 11): 'UAL',
    (1, 12): 'LAL',
    (1, 13): 'Outlier',
    (1, 14): 'Notes',
    (1, 15): 'Mean',
    (1, 16): 'Standard Deviation',
    (1, 17): 'Warning Limit',
    (1, 18): 'Action Limit',
    (1, 20): 'Q1',
    (1, 21): 'Q3',
    (1, 22): 'IQR',
    (1, 23): 'Upper Minor Outlier',
    (1, 24): 'Lower Minor Outlier',
    (1, 25): 'Upper Major Outlier',
    (1, 26): 'Lower Major Outlier',
    (1, 28): 'ETS Result',
    (1, 29): 'Previous LCS StDev',
}

STATS_DATA_VALUES = {
    (2, 15): '=AVERAGEIF(M:M, "No",D:D)',
    (2,
     16): '=IF(COUNT(Data[Measured Value])>20, STDEV(IF(M:M="No",D:D)), IF(ISBLANK(AC2),STDEV(IF(M:M="No",D:D)),$AC$2))',
    (2, 17): '=2*$P$2',
    (2, 18): '=3*$P$2',
    (2, 20): '=QUARTILE(D:D,1)',
    (2, 21): '=QUARTILE(D:D,3)',
    (2, 22): '=U2-T2',
    (2, 23): '=U2+(1.5*V2)',
    (2, 24): '=T2-(1.5*V2)',
    (2, 25): '=U2+(3*V2)',
    (2, 26): '=T2-(3*V2)',
}

DONT_TOUCH_VALUE = {
    (3, 15): '^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^ DON\'T TOUCH '
             '^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^',
}

DATA_TABLE_COLS_WITH_FORMULAS = {
    5: '=$O$2',
    6: '=IF(OR(NOT(ISNUMBER([@Time])), LEN([@Time]) > 4, LEN([@Time]) < 3, '
       'NOT(ISNUMBER([@Date])), LEN([@Date]) <>  5), DATE(2017, 5, 30) + TIME(24,0,0),'
       '[@Date]+TIMEVALUE(LEFT([@Time],LEN([@Time])-2)&":"&RIGHT([@Time],2)))',
    7: '=$O$2+$P$2',
    8: '=$O$2-$P$2',
    9: '=$O$2+$Q$2',
    10: '=$O$2-($Q$2)',
    11: '=$O$2+$R$2',
    12: '=$O$2-$R$2',
    13: '=IF(OR([@[Measured Value]]<$Z$2, [@[Measured Value]]>$Y$2), "Yes", "No")',
}

DATA_TABLE_COLS_WITH_USER_INPUT = {
    1: (datetime.datetime,),
    2: (int,),
    3: (str,),
    4: (int, float,),
}

DATA_TABLE_CONFIG = {
    'name': 'Data',
    'num_cols': 14,
    'data_min_row': 2,
}


def gen_cell(row_iter: Generator[Tuple[Union[Cell, Any]], None, None]) -> Any:
    for row in row_iter:
        if len(row) > 1:
            raise ValueError('Pass only row_iters with width 1')
        else:
            yield row[0]


def check_i_chart(excel_file_path: str, issues: Dict[str, List[str]]):
    check_ranges: List[Dict[Tuple[int, int], str]] = [HEADER_VALUES, STATS_DATA_VALUES, DONT_TOUCH_VALUE]
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.worksheets[WORKSHEET_INDEX]

    excel_file_path = os.path.basename(excel_file_path)

    # check text values in first row of data worksheet
    # check 'stats data' formulas in second row of data worksheet
    for cell_addresses_and_expected_values in check_ranges:
        for cell_address, expectation in cell_addresses_and_expected_values.items():
            try:
                if ws.cell(*cell_address).value != expectation:
                    issues[excel_file_path].append('%s: %s should be %s' % (
                        excel_file_path, ws.cell(*cell_address).coordinate, expectation
                    ))
            except AttributeError:
                pass  # TODO can't check value of merged cell

    # check number of tables on data worksheet
    if len(ws._tables) > 1:
        issues[excel_file_path].append('%s: %s tables found, 1 expected' % (excel_file_path, len(ws._tables)))

    data_table: openpyxl.worksheet.table.Table = None
    # check table name
    try:
        data_table = [table for table in ws._tables if table.name == DATA_TABLE_CONFIG['name']][0]
    except IndexError:
        issues[excel_file_path].append('%s: tables named %s expected, none found. Could not validate LCS entries' % (
            excel_file_path, DATA_TABLE_CONFIG['name']
        ))
        return

    # check number of cols
    if len(data_table.tableColumns) is not DATA_TABLE_CONFIG['num_cols']:
        issues[excel_file_path].append('%s: %s columns expected, %s found' % (
            excel_file_path, DATA_TABLE_CONFIG['num_cols'], len(data_table.tableColumns)
        ))

    # build list of iterators covering data in data_table
    data_table_range = range_boundaries(data_table.ref)
    data_table_range = {
        'min_col': None,
        'min_row': max(data_table_range[1], DATA_TABLE_CONFIG['data_min_row']),
        'max_col': None,
        'max_row': data_table_range[3],
    }
    data_table_iters = {}
    for col in itertools.chain(DATA_TABLE_COLS_WITH_USER_INPUT.keys(), DATA_TABLE_COLS_WITH_FORMULAS):
        data_table_range['min_col'] = data_table_range['max_col'] = col
        data_table_iters[col] = ws.iter_rows(**data_table_range)

    # check each values in each column
    for col_id, data_iter in data_table_iters.items():
        if col_id in DATA_TABLE_COLS_WITH_FORMULAS.keys():
            expected_formula = DATA_TABLE_COLS_WITH_FORMULAS[col_id]
            for cell in gen_cell(data_iter):
                if cell.value is expected_formula:
                    pass
                else:
                    issues[excel_file_path].append('%s: invalid formula in cell %s' % (
                        excel_file_path, cell.coordinate
                    ))
        if col_id in DATA_TABLE_COLS_WITH_USER_INPUT.keys():
            for cell in gen_cell(data_iter):
                expected_types = DATA_TABLE_COLS_WITH_USER_INPUT[col_id]
                if isinstance(cell.value, expected_types):
                    pass
                else:
                    issues[excel_file_path].append('%s: %s expected in cell %s, %s found' % (
                        excel_file_path, expected_types, cell.coordinate, cell.value
                    ))


if __name__ == '__main__':
    dirs = [
        r'F:\LabData\Lab\ISO 17025\Control Chart',
        r'F:\LabData\Lab\ISO 17025\Control Chart\HPLC Control Charts',
        r'F:\LabData\Lab\ISO 17025\Control Chart\Metals by ICP-OES Control Charts',
        r'F:\LabData\Lab\ISO 17025\Control Chart\Y15 Control Charts'
    ]
    files = [os.path.join(dir_, file) for dir_ in dirs for file in os.listdir(dir_)]

    includes = ('xlsx',)
    excludes = ('~', 'Nightly')
    issues = {}
    for file in files:
        if any(include in file for include in includes) and not any(exclude in file for exclude in excludes):
            issues[os.path.basename(file)] = []
            check_i_chart(file, issues)
